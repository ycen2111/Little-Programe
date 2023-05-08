from openpyxl import load_workbook

def file_check(type=None,add=None):
    if type=="FET":
        workbook=load_workbook(filename=str(add))
    if type=="FA":
        workbook=load_workbook(filename=str(add))
    if type=="SF":
        workbook=load_workbook(filename=str(add))
#workbook = load_workbook(filename=r'C:\Users\86135\Desktop\error angel cal.xlsx') #insert excel file
#print("Open Sheet ",workbook.sheetnames)

#sheet = workbook['Test'] #read sheet
#print(sheet.dimensions)

#cell=sheet['A1'] #single coordinante
#cell=sheet.cell(row=1,column=1) #same
#print(cell.value)

#cells=sheet['A1:C26'] #range of coordinantes
#cells=sheet['A'] #columns
#cells=sheet[1:265] #row
#for i in sheet: print(i)