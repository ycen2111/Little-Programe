#import head file
import sys #import system file
import easygui as gui #import gui lib
from openpyxl import load_workbook #load existed sheet
from openpyxl import Workbook #create new sheet
from openpyxl.styles import PatternFill #fill cell's background color
import datetime #get date

#-----------------------------------------------------------------

class Excel(object):
    def __init__(self):
        self.add="None"
        self.warning_list=[]
        self.key_words=[]
        self.start_row=2

    #open an excel and also open one sheet after selecting
    def open_excel_and_sheet(self):
        self.WB=load_workbook(filename=str(self.add)) #open workbook
        
        if len(self.WB.sheetnames)!=1:  #check if have multiple sheets
            self.sheet_name=gui.choicebox(msg="Multiple sheets founds. Please select one sheet below:",title="sheet selection",choices=self.WB.sheetnames)
            if self.sheet_name is None: #no choices
                self.add="None"
        else: #only have one sheet
            self.sheet_name=self.WB.sheetnames
            self.sheet_name=str(self.sheet_name)[2:len(self.sheet_name)-3] #remove [' '] inside original sheet name

        if self.sheet_name is not None: #Open seleted sheet and check head line
            self.sheet=self.WB[str(self.sheet_name)]
            self.max_row=self.sheet.max_row
            self.max_col=self.sheet.max_column
        
        print("Open file "+self.add+", "+self.sheet_name+"\n")

    #generate new workbook and create head line
    def gen_new_workbook(self):
        self.WB=Workbook()
        self.sheet=self.WB.active

        #create head line
        self.max_col=0
        self.max_row=1
        for i in self.key_words:
            self.max_col+=1 #shift column coordination
            self.sheet.cell(row=1,column=self.max_col,value=str(i))
            i=keep_letter_and_num(i).lower() #only keep number and letter
            exec("self.{}_ID = self.max_col".format(i)) #mark key words' column number. like: self.UserEmail_ID=5

    #enter key head words
    def set_key_words(self,keywords):
        for i in keywords:
            self.key_words.append(i)

    #get value of one row. head line mode will remove all " " , and turn to small letter
    def read_one_row(self,row,headLine):
        row_data=[]
        if headLine: #read head line
            for i in range(1,self.sheet.max_column+1):
                head_word=self.sheet.cell(row=row,column=i).value
                if head_word is not None:
                    head_word=keep_letter_and_num(head_word).lower() #only keep number and letter
                row_data.append(head_word)
        else:
            for i in range(1,self.sheet.max_column+1):
                row_data.append(self.sheet.cell(row=row,column=i).value)
        return row_data

    #check if have target word in head word
    def check_key_word(self,head_word,target_word):
        count=0
        for i in head_word:
            count=count+1
            if i==target_word:
                exec("self.{}_ID = count".format(i)) #mark key words' column number. like: self.UserEmail_ID=5
                return True #finds target
        print("Not found key word: "+str(target_word))
        return False

    #find target message. column id is the column need be find, target message is the string need be find
    def find_target_row(self,column_ID,target_msg):
        target_msg=str(target_msg)[:-9].lower()
        max_row=self.max_row+1
        for row_num in range(2,max_row):
            cell_msg=str(self.sheet.cell(row=row_num,column=column_ID).value)[:-9].lower()
            if cell_msg==target_msg: #find target row number by small letter form
                return row_num
        
        return None #not find target message

    # save excel
    def save_excel(self):
        date=datetime.datetime.now() #get system time
        save_add=gui.filesavebox(title="Save sheet",default="OpenAndFilled_"+date.strftime('%y%m%d')+".xlsx")
        if save_add is None:
            if gui.ynbox(msg="Warning. The worksheet has not been saved. Please press YES if you want to save the worksheet."):
                print("Repeat the saving\n")
                self.save_excel()
            else:
                pass
        else:
            print("Saving sheet in "+str(save_add))
            self.WB.save(str(save_add))
            

#-----------------------------------------------------------------------------
#remove marks but keep letters and numbers in string(using ACSII form)
def keep_letter_and_num(msg):
    res=""
    for i in msg:
        acs_num=ord(i)
        if (acs_num<58 and acs_num>47) or (acs_num<123 and acs_num>96) or (acs_num<91 and acs_num>64):
            res+=i
    return res

#rewritw FTE with new head line
def rewrite_FTE_in_New(FTE,New):
    for row in range(2,FTE.max_row+1): #rewrite sheet
        New.sheet.cell(row=row,column=New.pstnid_ID,value=FTE.sheet.cell(row=row,column=FTE.pstnid_ID).value)
        New.sheet.cell(row=row,column=New.prsnnm_ID,value=FTE.sheet.cell(row=row,column=FTE.prsnnm_ID).value)
        New.sheet.cell(row=row,column=New.pstndesc_ID,value=FTE.sheet.cell(row=row,column=FTE.pstndesc_ID).value)
        New.sheet.cell(row=row,column=New.workemail_ID,value=FTE.sheet.cell(row=row,column=FTE.workemail_ID).value)
        New.sheet.cell(row=row,column=New.bandcltgroup_ID,value=FTE.sheet.cell(row=row,column=FTE.bandcltgroup_ID).value)
        New.sheet.cell(row=row,column=New.supvnm_ID,value=FTE.sheet.cell(row=row,column=FTE.supvnm_ID).value)
        New.sheet.cell(row=row,column=New.suprvsrpstnid_ID,value=FTE.sheet.cell(row=row,column=FTE.suprvsrpstnid_ID).value)
        New.sheet.cell(row=row,column=New.funclareadesc_ID,value=FTE.sheet.cell(row=row,column=FTE.funclareadesc_ID).value)
        New.sheet.cell(row=row,column=New.lvl2tosorgchief_ID,value=FTE.sheet.cell(row=row,column=FTE.lvl2tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl3tosorgchief_ID,value=FTE.sheet.cell(row=row,column=FTE.lvl3tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl4tosorgchief_ID,value=FTE.sheet.cell(row=row,column=FTE.lvl4tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl5tosorgchief_ID,value=FTE.sheet.cell(row=row,column=FTE.lvl5tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl6tosorgchief_ID,value=FTE.sheet.cell(row=row,column=FTE.lvl6tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl7tosorgchief_ID,value=FTE.sheet.cell(row=row,column=FTE.lvl7tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl8tosorgchief_ID,value=FTE.sheet.cell(row=row,column=FTE.lvl8tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.pygrdgrpcd_ID,value=FTE.sheet.cell(row=row,column=FTE.pygrdgrpcd_ID).value)

        if row%8==0:
            print("Rewriting FTE row "+str(row)+"/"+str(FTE.max_row)+", "+str(int(row*100/FTE.max_row))+"%\r",end="",flush=True)
    New.max_row=New.sheet.max_row
    print("Rewriting FTE row "+str(row)+"/"+str(FTE.max_row)+", "+str(int(row*100/FTE.max_row)))

#combine part of cells from FA in New
def combine_FA_in_New(FA,New):
    FA.start_row= New.max_row
    for row in range(2,FA.max_row+1):
        New_row=New.max_row+row-1 #row number in New sheet
        New.sheet.cell(row=New_row,column=New.pstnid_ID,value=FA.sheet.cell(row=row,column=FA.empid_ID).value)
        New.sheet.cell(row=New_row,column=New.prsnnm_ID,value=FA.sheet.cell(row=row,column=FA.empenname_ID).value)
        New.sheet.cell(row=New_row,column=New.pstndesc_ID,value=FA.sheet.cell(row=row,column=FA.position_ID).value)
        New.sheet.cell(row=New_row,column=New.workemail_ID,value=FA.sheet.cell(row=row,column=FA.email_ID).value)
        New.sheet.cell(row=New_row,column=New.bandcltgroup_ID,value="FA ESP")
        New.sheet.cell(row=New_row,column=New.supvnm_ID,value=FA.sheet.cell(row=row,column=FA.enduser_ID).value)
        New.sheet.cell(row=New_row,column=New.construct_ID,value=row) #temp mark FA row value

        if row%8==0:
            print("Rewriting FA row "+str(row)+"/"+str(FA.max_row)+", "+str(int(row*100/FA.max_row))+"%\r",end="",flush=True)
    New.max_row=New.sheet.max_row
    print("Rewriting FA row "+str(row)+"/"+str(FA.max_row)+", "+str(int(row*100/FA.max_row))+"%\r")

#combine part of cells from SW in New
def combine_SF_in_New(SF,New):
    SF.start_row=New.max_row
    New_row=New.max_row
    for row in range(2,SF.max_row+1):
        type=SF.sheet.cell(row=row,column=SF.usertype_ID).value
        if type==4 or type==5: #only need type 4 and 5
            if New.find_target_row(New.workemail_ID,SF.sheet.cell(row=row,column=SF.useremail_ID).value) is None: #pass the repeat test
                New_row=New_row+1
                New.sheet.cell(row=New_row,column=New.pstnid_ID,value=SF.sheet.cell(row=row,column=SF.empid_ID).value)
                New.sheet.cell(row=New_row,column=New.prsnnm_ID,value=SF.sheet.cell(row=row,column=SF.username_ID).value)
                New.sheet.cell(row=New_row,column=New.pstndesc_ID,value=SF.sheet.cell(row=row,column=SF.userposition_ID).value)
                New.sheet.cell(row=New_row,column=New.workemail_ID,value=SF.sheet.cell(row=row,column=SF.useremail_ID).value)
                if type==4:
                    New.sheet.cell(row=New_row,column=New.bandcltgroup_ID,value="FA ESP")
                else: #type=5
                    New.sheet.cell(row=New_row,column=New.bandcltgroup_ID,value="Other ESP")
                New.sheet.cell(row=New_row,column=New.supvnm_ID,value=SF.sheet.cell(row=row,column=SF.linemanagername_ID).value)
                New.sheet.cell(row=New_row,column=New.company_ID,value=SF.sheet.cell(row=row,column=SF.company_ID).value)
                New.sheet.cell(row=New_row,column=New.construct_ID,value=row) #temp mark SW row value

            if row%8==0:
                print("Rewriting swoosh family row "+str(row)+"/"+str(SF.max_row)+", "+str(int(row*100/SF.max_row))+"%\r",end="",flush=True)
    New.max_row=New.sheet.max_row
    print("Rewriting swoosh family row "+str(row)+"/"+str(SF.max_row)+", "+str(int(row*100/SF.max_row))+"%\r")

def deal_manager_info(New,source,source_name):
    for row in range(source.start_row,New.max_row+1):
        if New.sheet.cell(row=row,column=New.suprvsrpstnid_ID).value is None: #havenot been filled
            find_and_fill_info(New,source,source_name,row)
        else:
            pass

        if row%8==0:
            if source_name=="FA":
                print("Finding FA manager in row "+str(row-FA.start_row)+"/"+str(FA.max_row)+", "+str(int((row-FA.start_row)*100/FA.max_row))+"%\r",end="",flush=True)
            elif source_name=="SF":
                print("Finding swoosh family manager in row "+str(row-SF.start_row)+"/"+str(New.max_row-SF.start_row)+", "+str(int((row-SF.start_row)*100/(New.max_row-SF.start_row)))+"%\r",end="",flush=True)
    
    for i in range(2,New.max_row+1): #remove temp id value
        New.sheet.cell(row=i,column=New.construct_ID,value="")

    if source_name=="FA":
        print("Finding FA manager in row "+str(FA.max_row)+"/"+str(FA.max_row)+", 100%\r")
    elif source_name=="SF":
        print("Finding swoosh family manager in row "+str(New.max_row-SF.start_row)+"/"+str(New.max_row-SF.start_row)+", 100%\r")

def find_and_fill_info(New,source,source_name,row):
    source_row=New.sheet.cell(row=row,column=New.construct_ID).value #re-coordinate row num in source sheet

    if source_name=="FA":
        manager_add=source.sheet.cell(row=source_row,column=source.euemailaddress_ID).value #find mamager address
    elif source_name=="SF":
        manager_add=source.sheet.cell(row=source_row,column=source.linemanageremail_ID).value #find mamager address
    else:
        print("warning!! new sheet type")

    manager_row=New.find_target_row(New.workemail_ID,manager_add) #get manager row in New sheet

    if manager_row is None: #not found manager info
        pass
        ##############################################
        ##dont need this type of waring list
        #warn="No manager found in "+source_name+" row "+str(source_row)
        #print(warn+"                    ")
        #New.warning_list.append(warn)
        ##############################################
    else: #found target manager
        if New.sheet.cell(row=manager_row,column=New.suprvsrpstnid_ID).value is None: #havenot been filled
            find_and_fill_info(New,source,source_name,manager_row)
        else:
            pass
        New.sheet.cell(row=row,column=New.suprvsrpstnid_ID,value=New.sheet.cell(row=manager_row,column=New.pstnid_ID).value)
        New.sheet.cell(row=row,column=New.lvl2tosorgchief_ID,value=New.sheet.cell(row=manager_row,column=New.lvl2tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl3tosorgchief_ID,value=New.sheet.cell(row=manager_row,column=New.lvl3tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl4tosorgchief_ID,value=New.sheet.cell(row=manager_row,column=New.lvl4tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl5tosorgchief_ID,value=New.sheet.cell(row=manager_row,column=New.lvl5tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl6tosorgchief_ID,value=New.sheet.cell(row=manager_row,column=New.lvl6tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl7tosorgchief_ID,value=New.sheet.cell(row=manager_row,column=New.lvl7tosorgchief_ID).value)
        New.sheet.cell(row=row,column=New.lvl8tosorgchief_ID,value=New.sheet.cell(row=manager_row,column=New.lvl8tosorgchief_ID).value)

#-----------------------------------------------------------------------------

#Start FSM
if __name__ == "__main__":
    #Variable initialization
    print("Start program\n\n")
    FTE=Excel()
    FA=Excel()
    SF=Excel()
    
    FTE.set_key_words(["pstnid","prsnnm","pstndesc","workemail","bandcltgroup","supvnm","suprvsrpstnid","funclareadesc","lvl2tosorgchief","lvl3tosorgchief","lvl4tosorgchief","lvl5tosorgchief","lvl6tosorgchief","lvl7tosorgchief","lvl8tosorgchief","pygrdgrpcd"])
    FA.set_key_words(["empid","empenname","enduser","euemailaddress","position","email"])
    SF.set_key_words(["empid","username","useremail","userposition","usertype","linemanagername","linemanageremail","company"])

    #create new workbook
    New=Excel() #new generated sheet
    New.set_key_words(["Pstn ID","Prsn Nm","Pstn Desc","Work Email","Band (CLT Group)","Supv Nm","SuprvsrPstnID","Construct","Funcl Area Desc","Lvl2TOSOrgChief","Lvl3TOSOrgChief","Lvl4TOSOrgChief","Lvl5TOSOrgChief","Lvl6TOSOrgChief","Lvl7TOSOrgChief","Lvl8TOSOrgChief","Py Grd Grp Cd","Company"])
    New.gen_new_workbook()

    #FSM definition
    INIT=0
    IMP_FTE=1
    IMP_FA=2
    IMP_SF=3
    STEP3=4
    MERGING=5
    FINISH=6
    EXIT=7

    state=INIT
    while 1:
        if state==INIT:  #Initialize window
            button=gui.buttonbox(msg="Welcome to Excel merging program.\n\n\n\
    Please import 3 documens in following order:\n\n\
        FTE: "\
        +str(FTE.add)+"\n\n\
        FA: "\
        +str(FA.add)+"\n\n\
        Swoosh family:  "\
        +str(SF.add)\
            ,title="Excel Merging",choices=["FTE","FA","Swoosh family","Submit"])#button box return string
            if button=="FTE":
                state=IMP_FTE
            elif button=="FA":
                state=IMP_FA
            elif button=="Swoosh family":
                state=IMP_SF
            elif button=="Submit":
                state=MERGING
            else:
                state=EXIT

        elif state==IMP_FTE: #import FTE
            FTE.add = gui.fileopenbox(title='open FTE file')
            if FTE.add is not None: #get address
                FTE.open_excel_and_sheet()

                for i in FTE.key_words: #enter kay words
                    if not FTE.check_key_word(FTE.read_one_row(1,True),i): #check key words
                        gui.msgbox(msg="Warning!!!\n\nCannot found at least one of following key words:\n"+str(FTE.key_words)+"\n\nPlease select Excel again",title="Open Tuser failed")
                        FTE.add="None"
                        break
            else:
                FTE.add="None"
            state=INIT

        elif state==IMP_FA: #import FA
            FA.add = gui.fileopenbox(title='open FA')
            if FA.add is not None: #get address
                FA.open_excel_and_sheet()

                for i in FA.key_words: #enter kay words
                    if not FA.check_key_word(FA.read_one_row(1,True),i): #check key words
                        gui.msgbox(msg="Warning!!!\n\nCannot found at least one of following key words:\n"+str(FA.key_words)+"\n\nPlease select Excel again",title="Open EE failed")
                        FA.add="None"
                        break
            else:
                FA.add="None"
            state=INIT

        elif state==IMP_SF: #import SF
            SF.add = gui.fileopenbox(title='open Swoosh Family')
            if SF.add is not None: #get address
                SF.open_excel_and_sheet()

                for i in SF.key_words: #enter kay words
                    if not SF.check_key_word(SF.read_one_row(1,True),i): #check key words
                        gui.msgbox(msg="Warning!!!\n\nCannot found at least one of following key words:\n"+str(SF.key_words)+"\n\nPlease select Excel again",title="Open Mapping failed")
                        SF.add="None"
                        break
            else:
                SF.add="None"
            state=INIT

        elif state==MERGING:
            if FTE.add=="None" or FA.add=="None" or SF.add=="None": #missing file
                missing=[]
                if FTE.add=="None":
                    missing.append("FTE")
                if FA.add=="None":
                    missing.append("FA")
                if SF.add=="None":
                    missing.append("SF")
                gui.msgbox(msg="Warning:\n\n\n\nMissing "+str(missing),title="Core file missing")
                state=FINISH
                
            else:
                rewrite_FTE_in_New(FTE,New) #read FTE staff list
                combine_FA_in_New(FA,New) #read FA staff list
                deal_manager_info(New,FA,"FA") #find staff manager in FA
                combine_SF_in_New(SF,New) #read SF staff list
                deal_manager_info(New,SF,"SF") #find staff manager in FA

                print("\nStart saving...\n")
                state=FINISH

        elif state==FINISH:
            New.save_excel()
            if len(New.warning_list)>0:
                text=""
                for i in New.warning_list: #list out warning content
                    text+=(str(i)+"\n")
                gui.textbox(msg="Found "+str(len(New.warning_list))+" warnings:",text=text,title="Warning list")
            else:
                gui.msgbox("Finished")
            break
        
        elif state==EXIT:
            break

        else: 
            print("Error: no state finding!!!") #alart no state finding
            state=FINISH

    sys.exit(0)