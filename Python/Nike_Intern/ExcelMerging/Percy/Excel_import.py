from openpyxl import load_workbook

#open new file
#return workbook
def open_excel(add=None):
    return load_workbook(filename=str(add))

#read sheet names in workbook
#return sheet names
def read_sheet_name(WB=None):
    return WB.sheetnames

def open_sheet(WB=None,sheet_name=None):
    name=str(sheet_name)[2:len(sheet_name)-3] #remove [' '] inside original sheet name

    return WB[str(name)]

def head_check(sheet=None,key_word=None):
    row=sheet["1"]
    return

#workbook = load_workbook(filename=r'C:\Users\86135\Desktop\error angel cal.xlsx') #insert excel file
#print("Open Sheet ",workbook.sheetnames)

#sheet = workbook['Test'] #read sheet
#print(sheet.dimensions)

#cell=sheet['A1'] #single coordinante
#cell=sheet.cell(row=1,column=1) #same
#print(cell.value)

#cells=sheet['A1:C26'] #range of coordinantes
#cells=sheet['A'] #columns
#cells=sheet[1:265] #row
#for i in sheet: print(i)