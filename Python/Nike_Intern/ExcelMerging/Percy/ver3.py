#import head file
import sys #import system file
import easygui as gui #import gui lib
from openpyxl import load_workbook

#-----------------------------------------------------------------

class Excel(object):
    def __init__(self):
        self.add="None"
        self.warning_list=[]
        self.key_words=[]

    #open an excel and also open one sheet after selecting
    def open_excel_and_sheet(self):
        self.WB=load_workbook(filename=str(self.add)) #open workbook
        if len(self.WB.sheetnames)!=1:  #check if have multiple sheets
            self.sheet_name=gui.choicebox(msg="Find multiple sheets. Please select one sheet below:",title="sheet selection",choices=self.WB.sheetnames)
            if self.sheet_name is None: #no choices
                self.add="None"
        else: #only have one sheet
            self.sheet_name=self.WB.sheetnames
            self.sheet_name=str(self.sheet_name)[2:len(self.sheet_name)-3] #remove [' '] inside original sheet name

        if self.sheet_name is not None: #Open seleted sheet and check head line
            self.sheet=self.WB[str(self.sheet_name)]
            self.max_row=self.sheet.max_row
            self.max_col=self.sheet.max_column

    #enter key head words
    def set_key_words(self,keywords):
        for i in keywords:
            self.key_words.append(i)

    #get value of one row. head line mode will remove all " " , and turn to small letter
    def read_one_row(self,row,headLine):
        row_data=[]
        if headLine: #read head line
            for i in range(1,self.sheet.max_column):
                head_word=self.sheet.cell(row=row,column=i).value
                if head_word is not None:
                    head_word=head_word.replace(" ","").lower() #remove " "
                row_data.append(head_word)
        else:
            for i in range(1,self.sheet.max_column):
                row_data.append(self.sheet.cell(row=row,column=i).value)
        return row_data

    #check if have target word in head word
    def check_key_word(self,head_word,target_word):
        count=0
        for i in head_word:
            count=count+1
            if i==target_word:
                exec("self.{}_ID = count".format(i)) #mark key words' column number. like: self.UserEmail_ID=5
                return True #finds target
        return False

    #find target message. column id is the column need be find, target message is the string need be find
    def find_target_row(self,column_ID,target_msg):
        target_msg=str(target_msg)[:-9].lower()
        max_row=self.max_row+1
        for row_num in range(2,max_row):
            cell_msg=str(self.sheet.cell(row=row_num,column=column_ID).value)[:-9].lower()
            if cell_msg==target_msg: #find target row number by small letter form
                return row_num
        
        return None #not find target message

    # save excel
    def save_excel(self):
        self.WB.save("./test.xlsx")

#-----------------------------------------------------------------------------
#@profile  
def Userdepartment_ID_enter(Tuser,EE,Mapping):
    max_row=Tuser.sheet.max_row
    max_column=Tuser.sheet.max_column+1

    for row_num in range(2,max_row+1):
        if str(Tuser.sheet.cell(row=row_num,column=Tuser.userdepartmentid_ID).value)=="0":
                    Tuser.sheet.cell(row=row_num,column=Tuser.userdepartmentid_ID,value=1356) #turn 0 as 1356

    for row_num in range(2,max_row+1):
        if str(Tuser.sheet.cell(row=row_num,column=Tuser.usertype_ID).value)=="3": #gap if usertype=3
            Tuser.sheet.cell(row=row_num,column=max_column,value=1) #mark finish sign
            #print("Ignor row "+str(row_num))

        if Tuser.sheet.cell(row=row_num,column=max_column).value!=1: #havenot been serched
            #print("row"+str(row_num))
            Find_Manager_Id(Tuser,EE,Mapping,row_num)
        

#Recursion. find core manager department ID    
def Find_Manager_Id(Tuser,EE,Mapping,row_num):
    final_id=None
        
    #if str(Tuser.sheet.cell(row=row_num,column=Tuser.sheet.max_column).value)!="1": #havent been serched
    Manager_row=Tuser.find_target_row(Tuser.useremail_ID,Tuser.sheet.cell(row=row_num,column=Tuser.linemanageremail_ID).value) #find out manager row number
    #print(Manager_row)

    if Manager_row is None: #find core manager row
        EE_Staff_row=EE.find_target_row(EE.workemail_ID,Tuser.sheet.cell(row=row_num,column=Tuser.useremail_ID).value) #find user row in EE form
        if EE_Staff_row is not None: #find staff row in EE
            Staff_function=str(EE.sheet.cell(row=EE_Staff_row,column=EE.function3_ID).value).lower()
            for i in range(2,Mapping.max_row+1):
                if str(Mapping.sheet.cell(row=i,column=Mapping.function3_ID).value).lower()==Staff_function: #find function 3 in mapping sheet
                    final_id=Mapping.sheet.cell(row=i,column=Mapping.departmentid_ID).value
                    break

            if final_id is None: #not find department in Mapping sheet
                final_id=1356
                print("warning cannot find "+EE.sheet.cell(row=EE_Staff_row,column=EE.function3_ID).value) #not find department in mapping sheet
        else: #not find staff mail in EE sheet
            final_id=1356
            print("warning cannot find row "+str(row_num)+" staff in EE") #not find staff mail in EE sheet

    else: #find Manager's manager
        if Tuser.sheet.cell(row=Manager_row,column=Tuser.max_col+1).value==1: #manager have been searched
            #print("find finish mark in row "+str(Manager_row))
            final_id=Tuser.sheet.cell(row=Manager_row,column=Tuser.userdepartmentid_ID).value
        else: #new find manager row
            final_id=Find_Manager_Id(Tuser,EE,Mapping,Manager_row)
    
    #print(final_id)
    Tuser.sheet.cell(row=row_num,column=Tuser.userdepartmentid_ID,value=final_id) #write department id
    Tuser.sheet.cell(row=row_num,column=Tuser.max_col+1,value=1) #mark finish sign
    return final_id

#-----------------------------------------------------------------------------

#Start FSM
if __name__ == "__main__":
    #Variable initialization
    Tuser=Excel()
    EE=Excel()
    Mapping=Excel()
    Tuser.set_key_words(["useremail","userdepartmentid","usertype","linemanageremail"])
    EE.set_key_words(["function3","workemail"])
    Mapping.set_key_words(["function3","departmentid"])

    #FSM definition
    INIT=0
    IMP_TUSER=1
    IMP_EE=2
    IMP_MAPPING=3
    STEP3=4
    MERGING=5
    FINISH=6
    EXIT=7

    state=INIT
    while 1:
        if state==INIT:  #Initialize window
            button=gui.buttonbox(msg="Welcome to Excel merging program.\n\n\n\
    Please import 3 documens in following order:\n\n\
        Tuser: "\
        +str(Tuser.add)+"\n\n\
        FTE+FA ESP: "\
        +str(EE.add)+"\n\n\
        Mapping:  "\
        +str(Mapping.add)\
            ,title="Excel Merging",choices=["Tuser","FTE+FA ESP","MAPPING","Submit"])#button box return string
            if button=="Tuser":
                state=IMP_TUSER
            elif button=="FTE+FA ESP":
                state=IMP_EE
            elif button=="MAPPING":
                state=IMP_MAPPING
            elif button=="Submit":
                state=MERGING
            else:
                state=EXIT

        elif state==IMP_TUSER: #import Tuser
            Tuser.add = gui.fileopenbox(title='open Tuser file')
            if Tuser.add is not None: #get address
                Tuser.open_excel_and_sheet()

                for i in Tuser.key_words: #enter kay words
                    if not Tuser.check_key_word(Tuser.read_one_row(1,True),i): #check key words
                        gui.msgbox(msg="Warning!!!\n\nCannot find at least one of following key words:\n"+str(Tuser.key_words)+"\n\nPlease select Excel again",title="Open Tuser failed")
                        Tuser.add="None"
                        break
            else:
                Tuser.add="None"
            state=INIT

        elif state==IMP_EE: #import FTE, FA
            EE.add = gui.fileopenbox(title='open FTE+FA ESP')
            if EE.add is not None: #get address
                EE.open_excel_and_sheet()

                for i in EE.key_words: #enter kay words
                    if not EE.check_key_word(EE.read_one_row(1,True),i): #check key words
                        gui.msgbox(msg="Warning!!!\n\nCannot find at least one of following key words:\n"+str(EE.key_words)+"\n\nPlease select Excel again",title="Open EE failed")
                        EE.add="None"
                        break
            else:
                EE.add="None"
            state=INIT

        elif state==IMP_MAPPING: #import Mapping
            Mapping.add = gui.fileopenbox(title='open Mapping')
            if Mapping.add is not None: #get address
                Mapping.open_excel_and_sheet()

                for i in Mapping.key_words: #enter kay words
                    if not Mapping.check_key_word(Mapping.read_one_row(1,True),i): #check key words
                        gui.msgbox(msg="Warning!!!\n\nCannot find at least one of following key words:\n"+str(Mapping.key_words)+"\n\nPlease select Excel again",title="Open Mapping failed")
                        Mapping.add="None"
                        break
            else:
                Mapping.add="None"
            state=INIT

        elif state==MERGING:
            if Tuser.add=="None" or EE.add=="None" or Mapping.add=="None": #missing file
                missing=[]
                if Tuser.add=="None":
                    missing.append("Tuser")
                if EE.add=="None":
                    missing.append("FTE and ESP")
                if Mapping.add=="None":
                    missing.append("Mapping")
                gui.msgbox(msg="Warning:\n\n\n\nMissing "+str(missing),title="Core file missing")
                state=INIT
            else:
                Tuser.sheet.cell(row=1,column=Tuser.sheet.max_column+1,value=1) #mark sign

                #Find_Manager_Id(Tuser,EE,Mapping,22)
                #Find_Manager_Id(Tuser,EE,Mapping,1646)
                Userdepartment_ID_enter(Tuser,EE,Mapping)

                state=FINISH

        elif state==FINISH:
            Tuser.save_excel()
            gui.msgbox("Finished")
            break
        
        elif state==EXIT:
            break

        else: 
            print("Error: no state finding!!!") #alart no state finding
            state=FINISH

    sys.exit(0)