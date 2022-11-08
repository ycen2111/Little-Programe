#pack code: pyinstaller --name="test" --console --onefile --icon="C:\Users\86135\Desktop\Intern\Percy\icons.ico" C:\Users\86135\Desktop\Intern\Percy\fin_ver_2.py 

#import head file
import sys #import system file
import easygui as gui #import gui lib
from openpyxl import load_workbook
from openpyxl import Workbook #create new sheet

#-----------------------------------------------------------------

class Excel(object):
    def __init__(self):
        self.add="None"
        self.warning_list=[]
        self.key_words=[]

    #open an excel and also open one sheet after selecting
    def open_excel_and_sheet(self):
        self.WB=load_workbook(filename=str(self.add)) #open workbook
        
        if len(self.WB.sheetnames)!=1:  #check if have multiple sheets
            self.sheet_name=gui.choicebox(msg="Multiple sheets found. Please select one sheet below:",title="sheet selection",choices=self.WB.sheetnames)
            if self.sheet_name is None: #no choices
                self.add="None"
        else: #only have one sheet
            self.sheet_name=self.WB.sheetnames
            self.sheet_name=str(self.sheet_name)[2:len(self.sheet_name)-3] #remove [' '] inside original sheet name

        if self.sheet_name is not None: #Open seleted sheet and check head line
            self.sheet=self.WB[str(self.sheet_name)]
            self.max_row=self.sheet.max_row
            self.max_col=self.sheet.max_column
        
        if self.sheet_name is not None:
            print("Open file "+self.add+", "+self.sheet_name+"\n")

    #generate new sheet and create head line
    def gen_new_sheet(self,WB,sheet_id,name):
        self.WB=WB
        self.sheet=WB.create_sheet(str(name),sheet_id)

        #create head line
        self.max_col=0
        self.max_row=1
        for i in self.key_words:
            self.max_col+=1 #shift column coordination
            self.sheet.cell(row=1,column=self.max_col,value=str(i))
            i=keep_letter_and_num(str(i)).lower() #only keep number and letter
            exec("self.{}_ID = self.max_col".format(i)) #mark key words' column number. like: self.UserEmail_ID=5

    #enter key head words
    def set_key_words(self,keywords):
        for i in keywords:
            self.key_words.append(i)

    #get value of one row. head line mode will remove all " " , and turn to small letter
    def read_one_row(self,row,headLine):
        row_data=[]
        if headLine: #read head line
            for i in range(1,self.sheet.max_column+1):
                head_word=self.sheet.cell(row=row,column=i).value
                if head_word is not None:
                    head_word=keep_letter_and_num(head_word).lower() #only keep number and letter
                row_data.append(head_word)
        else:
            for i in range(1,self.sheet.max_column+1):
                row_data.append(self.sheet.cell(row=row,column=i).value)
        return row_data

    #check if have target word in head word
    def check_key_word(self,head_word,target_word):
        count=0
        for i in head_word:
            count=count+1
            if i==target_word:
                exec("self.{}_ID = count".format(i)) #mark key words' column number. like: self.UserEmail_ID=5
                return True #finds target
        print("Not found key word: "+str(target_word))
        return False

    #find target message. column id is the column need be find, target message is the string need be find
    def find_target_row(self,column_ID,target_msg):
        target_msg=str(target_msg)[:-9].lower()
        for row_num in range(2,self.max_row+1):
            cell_msg=str(self.sheet.cell(row=row_num,column=column_ID).value)[:-9].lower()
            if cell_msg==target_msg: #find target row number by small letter form
                return row_num
        
        return None #not find target message

    # save excel
    def save_excel(self):
        save_add=gui.filesavebox(title="Save sheet",default=str(self.add)[:-5]+"_copy.xlsx")
        if save_add is None:
            if gui.ynbox(msg="Warning. The worksheet has not been saved. Please press YES if you want to save the worksheet."):
                print("Repeat the saving\n")
                self.save_excel()
            else:
                pass
        else:
            print("Saving sheet in "+str(save_add))
            self.WB.save(str(save_add))
            

#-----------------------------------------------------------------------------
#import and self-check the sheet
def import_sheet(name,sheet):
    sheet.add = gui.fileopenbox(title='open '+name+' file')
    if sheet.add is not None: #get address
        sheet.open_excel_and_sheet()

        if sheet.sheet_name is not None:
            for i in sheet.key_words: #enter kay words
                if not sheet.check_key_word(sheet.read_one_row(1,True),i): #check key words
                    gui.msgbox(msg="Warning!!!\n\nCannot found at least one of following key words:\n"+str(sheet.key_words)+"\n\nPlease select Excel again",title="Open "+name+" failed")
                    sheet.add="None"
                    break
    else:
        sheet.add="None"

#copy pre sheet in new sheet
def copy_sheet(pre_sheet,new_sheet):
    row_num=0
    for row_value in pre_sheet.sheet.rows:
        column_num=0
        row_num=row_num+1
        for cell in row_value:
            column_num=column_num+1
            new_sheet.sheet.cell(row=row_num,column=column_num,value=cell.value)

#remove marks but keep letters and numbers in string(using ACSII form)
def keep_letter_and_num(msg):
    res=""
    for i in msg:
        acs_num=ord(i)
        if (acs_num<58 and acs_num>47) or (acs_num<123 and acs_num>96) or (acs_num<91 and acs_num>64):
            res+=i
    return res

#@profile  
count=0
def Userdepartment_ID_enter(Tuser,EE,Mapping):
    max_row=Tuser.sheet.max_row
    max_column=Tuser.sheet.max_column+1
    global count #process counting

    for row_num in range(2,max_row+1):
        if str(Tuser.sheet.cell(row=row_num,column=Tuser.userdepartmentid_ID).value)=="0":
                    Tuser.sheet.cell(row=row_num,column=Tuser.userdepartmentid_ID,value=1356) #turn 0 as 1356

    for row_num in range(2,max_row+1):
        if str(Tuser.sheet.cell(row=row_num,column=Tuser.usertype_ID).value)=="3": #gap if usertype=3
            Tuser.sheet.cell(row=row_num,column=max_column,value=1) #mark finish sign
            count+=1
            #print("Ignor row "+str(row_num))

        if Tuser.sheet.cell(row=row_num,column=max_column).value!=1: #havenot been serched
            #print("row"+str(row_num))
            Find_Manager_Id(Tuser,EE,Mapping,row_num)
        
        if not row_num%10:
            per=int(count*100/max_row)
            if (per<101):
                print("Finding row "+str(count)+"/"+str(max_row)+", "+str(per)+"%\r",end="",flush=True)
        

#Recursion. find manager department ID and check current position id  
def Find_Manager_Id(Tuser,EE,Mapping,row_num):
    global count
    final_id=None
    
    if Tuser.sheet.cell(row=row_num,column=Tuser.userdepartmentid_ID).value==1356: #void department id
        manager_row=Tuser.find_target_row(Tuser.useremail_ID,Tuser.sheet.cell(row=row_num,column=Tuser.linemanageremail_ID).value) #find linemanager row
        if manager_row is not None: #found manager row
            final_id=Find_Manager_Id(Tuser,EE,Mapping,manager_row)
        else:
            pass
    else: #has department id. check it
        EE_row=EE.find_target_row(EE.workemail_ID,Tuser.sheet.cell(row=row_num,column=Tuser.useremail_ID).value)
        if EE_row is not None: #has staff list in EE sheet
            Staff_function=str(EE.sheet.cell(row=EE_row,column=EE.function3_ID).value).lower()
            for i in range(2,Mapping.max_row+1):
                if str(Mapping.sheet.cell(row=i,column=Mapping.function3_ID).value).lower()==Staff_function: #find function 3 in mapping sheet
                    final_id=Mapping.sheet.cell(row=i,column=Mapping.departmentid_ID).value
                    break

            if final_id is None: #new department id, print a warning list
                final_id=1356
                warn="New department "+str(EE.sheet.cell(row=EE_row,column=EE.function3_ID).value)+" in row "+str(row_num) #write new department warning list
                Tuser.warning_list.append(warn)
                print(warn)
            else: #found delartment id
                pass
        else:
            final_id=1356 #no staff row is found in EE sheet
    
    #print(final_id)
    Tuser.sheet.cell(row=row_num,column=Tuser.userdepartmentid_ID,value=final_id) #write department id
    Tuser.sheet.cell(row=row_num,column=Tuser.max_col+1,value=1) #mark finish sign
    count+=1
    return final_id

#-----------------------------------------------------------------------------

#Start FSM
if __name__ == "__main__":
    #Variable initialization
    print("Start program\n\n")
    Tuser=Excel()
    EE=Excel()
    Mapping=Excel()
    Tuser.set_key_words(["useremail","userdepartmentid","usertype","linemanageremail"])
    EE.set_key_words(["function3","workemail"])
    Mapping.set_key_words(["function3","departmentid"])

    #FSM definition
    INIT=0
    IMP_TUSER=1
    IMP_EE=2
    IMP_MAPPING=3
    STEP3=4
    MERGING=5
    FINISH=6
    EXIT=7

    state=INIT
    while 1:
        if state==INIT:  #Initialize window
            button=gui.buttonbox(msg="Welcome to Excel merging program.\n\n\n\
    Please import 3 documens in following order:\n\n\
        Tuser: "\
        +str(Tuser.add)+"\n\n\
        FTE+FA ESP: "\
        +str(EE.add)+"\n\n\
        Mapping:  "\
        +str(Mapping.add)\
            ,title="Excel Merging",choices=["Tuser","FTE+FA ESP","Mapping","Submit"])#button box return string
            if button=="Tuser":
                import_sheet(button,Tuser)
            elif button=="FTE+FA ESP":
                import_sheet("EE",EE)
            elif button=="Mapping":
                import_sheet(button,Mapping)
            elif button=="Submit":
                state=MERGING
            else:
                state=EXIT

        elif state==IMP_TUSER: #import Tuser
            Tuser.add = gui.fileopenbox(title='open Tuser file')
            if Tuser.add is not None: #get address
                Tuser.open_excel_and_sheet()

                for i in Tuser.key_words: #enter kay words
                    if not Tuser.check_key_word(Tuser.read_one_row(1,True),i): #check key words
                        gui.msgbox(msg="Warning!!!\n\nCannot found at least one of following key words:\n"+str(Tuser.key_words)+"\n\nPlease select Excel again",title="Open Tuser failed")
                        Tuser.add="None"
                        break
            else:
                Tuser.add="None"
            state=INIT

        elif state==IMP_EE: #import FTE, FA
            EE.add = gui.fileopenbox(title='open FTE+FA ESP')
            if EE.add is not None: #get address
                EE.open_excel_and_sheet()

                for i in EE.key_words: #enter kay words
                    if not EE.check_key_word(EE.read_one_row(1,True),i): #check key words
                        gui.msgbox(msg="Warning!!!\n\nCannot found at least one of following key words:\n"+str(EE.key_words)+"\n\nPlease select Excel again",title="Open EE failed")
                        EE.add="None"
                        break
            else:
                EE.add="None"
            state=INIT

        elif state==IMP_MAPPING: #import Mapping
            Mapping.add = gui.fileopenbox(title='open Mapping')
            if Mapping.add is not None: #get address
                Mapping.open_excel_and_sheet()

                for i in Mapping.key_words: #enter kay words
                    if not Mapping.check_key_word(Mapping.read_one_row(1,True),i): #check key words
                        gui.msgbox(msg="Warning!!!\n\nCannot found at least one of following key words:\n"+str(Mapping.key_words)+"\n\nPlease select Excel again",title="Open Mapping failed")
                        Mapping.add="None"
                        break
            else:
                Mapping.add="None"
            state=INIT

        elif state==MERGING:
            if Tuser.add=="None" or EE.add=="None" or Mapping.add=="None": #missing file
                missing=[]
                if Tuser.add=="None":
                    missing.append("Tuser")
                if EE.add=="None":
                    missing.append("FTE and ESP")
                if Mapping.add=="None":
                    missing.append("Mapping")
                gui.msgbox(msg="Warning:\n\n\n\nMissing "+str(missing),title="Core file missing")
                state=INIT
            else:
                WB=Workbook()
                #create new workbook
                New=Excel() #new generated sheet
                New.set_key_words(Tuser.read_one_row(1,False))
                New.gen_new_sheet(WB=WB,sheet_id=0,name="Main")
                New.max_row=Tuser.max_row
                New.max_col=Tuser.max_col
                New.add=Tuser.add
                copy_sheet(Tuser,New)

                New.sheet.cell(row=1,column=Tuser.sheet.max_column+1,value=1) #mark sign
                Userdepartment_ID_enter(New,EE,Mapping)
                print("\nStart saving...\n")

                state=FINISH

        elif state==FINISH:
            New.save_excel()
            if len(New.warning_list)>0:
                text=""
                for i in New.warning_list: #list out warning content
                    text+=(str(i)+"\n")
                gui.textbox(msg="Found "+str(len(New.warning_list))+" warnings:",text=text,title="Warning list")
            else:
                gui.msgbox("Finished")
            break
        
        elif state==EXIT:
            break

        else: 
            print("Error: no state finding!!!") #alart no state finding
            state=FINISH

    sys.exit(0)