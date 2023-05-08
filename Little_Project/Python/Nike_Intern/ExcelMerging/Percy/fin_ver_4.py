#pack code: pyinstaller --name="ForPercy" --console --onefile --icon="D:\Python\Intern\ExcelMerging\Percy\icons.ico" --distpath="D:\Python\Intern\ExcelMerging\Percy" D:\Python\Intern\ExcelMerging\Percy\fin_ver_2.py 

#import head file
from glob import glob
import sys #import system file
import easygui as gui #import gui lib
from openpyxl import load_workbook
from openpyxl import Workbook

#-----------------------------------------------------------------

class Excel(object):
    def __init__(self):
        self.add="None"
        self.warning_list=[]
        self.key_words=[]

    #open an excel and also open one sheet after selecting
    def open_excel_and_sheet(self):
        self.WB=load_workbook(filename=str(self.add),data_only=True) #open workbook
        
        if len(self.WB.sheetnames)!=1:  #check if have multiple sheets
            self.sheet_name=gui.choicebox(msg="Multiple sheets found. Please select one sheet below:",title="sheet selection",choices=self.WB.sheetnames)
            if self.sheet_name is None: #no choices
                self.add="None"
        else: #only have one sheet
            self.sheet_name=self.WB.sheetnames
            self.sheet_name=str(self.sheet_name)[2:len(self.sheet_name)-3] #remove [' '] inside original sheet name

        if self.sheet_name is not None: #Open seleted sheet and check head line
            self.sheet=self.WB[str(self.sheet_name)]
            self.max_row=self.sheet.max_row
            self.max_col=self.sheet.max_column
        
        if self.sheet_name is not None:
            print("Open file "+self.add+", "+self.sheet_name+"\n")

    #generate new sheet and create head line
    def gen_new_sheet(self,WB,sheet_id,name):
        self.WB=WB

        if len(WB.sheetnames)==1:
            self.sheet=WB.active
        else:
            self.sheet=WB.create_sheet(str(name),sheet_id)

        #create head line
        self.max_col=0
        self.max_row=1
        for i in self.key_words:
            self.max_col+=1 #shift column coordination
            self.sheet.cell(row=1,column=self.max_col,value=str(i))
            i=keep_letter_and_num(str(i)).lower() #only keep number and letter
            exec("self.{}_ID = self.max_col".format(i)) #mark key words' column number. like: self.UserEmail_ID=5

    #enter key head words
    def set_key_words(self,keywords):
        for i in keywords:
            self.key_words.append(i)

    #get value of one row. head line mode will remove all " " , and turn to small letter
    def read_one_row(self,row,headLine):
        row_data=[]
        if headLine: #read head line
            for i in range(1,self.sheet.max_column+1):
                head_word=self.sheet.cell(row=row,column=i).value
                if head_word is not None:
                    head_word=keep_letter_and_num(head_word).lower() #only keep number and letter
                row_data.append(head_word)
        else:
            for i in range(1,self.sheet.max_column+1):
                row_data.append(self.sheet.cell(row=row,column=i).value)
        return row_data

    #check if have target word in head word
    def check_key_word(self,head_word,target_word):
        count=0
        for i in head_word:
            count=count+1
            if i==target_word:
                exec("self.{}_ID = count".format(i)) #mark key words' column number. like: self.UserEmail_ID=5
                return True #finds target
        print("Not found key word: "+str(target_word))
        return False

    #create new dictionary
    def create_dict(self,dict_name):
        exec("self.{}_dict={}".format(dict_name,{})) #create new dictionary. like self.email_dict={}

    # save excel
    def save_excel(self):
        save_add=gui.filesavebox(title="Save sheet",default=str(self.add)[:-5]+"_copy")
        save_add=save_add+".xlsx"
        if save_add is None:
            if gui.ynbox(msg="Warning. The worksheet has not been saved. Please press YES if you want to save the worksheet."):
                print("Repeat the saving\n")
                self.save_excel()
            else:
                pass
        else:
            print("Saving sheet in "+str(save_add))
            self.WB.save(str(save_add))
            

#-----------------------------------------------------------------------------
#import and self-check the sheet
def import_sheet(name,sheet):
    sheet.add = gui.fileopenbox(title='open '+name+' file')
    if sheet.add is not None: #get address
        sheet.open_excel_and_sheet()

        if sheet.sheet_name is not None:
            for i in sheet.key_words: #enter kay words
                if not sheet.check_key_word(sheet.read_one_row(1,True),i): #check key words
                    gui.msgbox(msg="Warning!!!\n\nCannot found at least one of following key words:\n"+str(sheet.key_words)+"\n\nPlease select Excel again",title="Open "+name+" failed")
                    sheet.add="None"
                    break
    else:
        sheet.add="None"

#append dictionart
#input:
#   dict: searched dictionary
#   key: key
#   word: word
#output:
#   None
def append_dict(dict,key,word,list_type=False):
    if list_type is True:
        dict[str(key).upper()]=[word[0]]
        if len(word)>1:
            for i in word[1:]:
                dict[str(key).upper()].append(i)
    else:
        dict[str(key).upper()]=word


#search dictionary by key
#input:
#   dict: searched dictionary
#   key: key
#output:
#   word or None
def search_dict(dict,key):
    if str(key).upper() in dict:
        return dict[str(key).upper()]
    else:
        return None

#copy pre sheet in new sheet
def copy_sheet(pre_sheet,new_sheet):
    row_num=0
    new_sheet.max_row=pre_sheet.max_row
    new_sheet.max_col=pre_sheet.max_col
    new_sheet.add=pre_sheet.add

    row=0
    for row_value in pre_sheet.sheet.rows:
        row=row+1
        column_num=0
        row_num=row_num+1
        for cell in row_value:
            column_num=column_num+1
            new_sheet.sheet.cell(row=row_num,column=column_num,value=cell.value)
        
        if not row%8:
            per=int(row*100/pre_sheet.max_row)
            if (per<101):
                print("Copying row "+str(row)+"/"+str(pre_sheet.max_row)+", "+str(per)+"%\r",end="",flush=True)
            else:
                print("Copying row "+str(pre_sheet.max_row)+"/"+str(pre_sheet.max_row)+", 100%\r")

#remove marks but keep letters and numbers in string(using ACSII form)
def keep_letter_and_num(msg):
    res=""
    for i in msg:
        acs_num=ord(i)
        if (acs_num<58 and acs_num>47) or (acs_num<123 and acs_num>96) or (acs_num<91 and acs_num>64):
            res+=i
    return res

count=0
def Search_EE_Sheet(New,EE,Mapping):
    global count

    for row in range(2,New.max_row+1):
        if New.sheet.cell(row=row,column=New.userdepartmentid_ID).value==0:
            New.sheet.cell(row=row,column=New.userdepartmentid_ID,value=1356)

    for row in range(2,New.max_row+1):
        if str(New.sheet.cell(row=row,column=New.usertype_ID).value)=="3": #gap if usertype=3
            New.sheet.cell(row=row,column=New.max_col+1,value=1) #mark finish sign
            count+=1
        else:
            if New.sheet.cell(row=row,column=New.max_col).value!=1: #havenot been serched
                EE_row=search_dict(EE.workemail_dict,New.sheet.cell(row=row,column=New.useremail_ID).value)
                if EE_row is not None: #found staff in EE
                    final_id=search_dict(Mapping.function3_dict,EE.sheet.cell(row=EE_row,column=EE.function3_ID).value) #find function id in mapping sheet
                    if final_id is not None: #got department id
                        New.sheet.cell(row=row,column=New.userdepartmentid_ID,value=final_id) #print department id
                    else: #new department description
                        final_id=0
                        warn="New department "+str(EE.sheet.cell(row=EE_row,column=EE.function3_ID).value)+" in row "+str(row) #write new department warning list
                        New.warning_list.append(warn)
                        print(warn)
                    New.sheet.cell(row=row,column=New.max_col+1,value=1) #mark finish sign
                    count+=1
                else: #not found in EE sheet, move to next step
                    pass

        if not row%8:
            per=int(count*100/New.max_row)
            if (per<101):
                print("Filled row "+str(count)+"/"+str(New.max_row)+", "+str(per)+"%\r",end="",flush=True)

    print("Matched "+str(count)+" rows from EE sheet")

def Find_Id_by_Manager(New):
    global count
    beginning_count=count

    for row in range(2,New.max_row+1):
        if New.sheet.cell(row=row,column=New.max_col+1).value!=1: #havenot been serched
            Find_Manager_ID(New,row)

        if not row%8:
            per=int(count*100/New.max_row)
            if (per<101):
                print("Filled row "+str(count)+"/"+str(New.max_row)+", "+str(per)+"%\r",end="",flush=True)

    print("Filled "+str(count-beginning_count)+" rows by the Manager column")
    return count-beginning_count

#Recursion. find manager department ID
def Find_Manager_ID(New,row):
    global count

    Manager_add=New.sheet.cell(row=row,column=New.linemanageremail_ID).value
    if Manager_add is None: #dont have manager
        pass
    else: 
        Manager_row=search_dict(New.useremail_dict,Manager_add) #find linemanager row
        if Manager_row is None: #dont have manager
            pass
        else:
            if New.sheet.cell(row=Manager_row,column=New.max_col+1).value==1: #Manager has confirmed department id
                final_id=New.sheet.cell(row=Manager_row,column=New.userdepartmentid_ID).value
            else:
                final_id=Find_Manager_ID(New,Manager_row)

            if final_id is not None: #exactly has final id
                count=count+1
                New.sheet.cell(row=row,column=New.max_col+1,value=1) #mark finish sign
                New.sheet.cell(row=row,column=New.userdepartmentid_ID,value=final_id) #print department id
                return final_id

def Find_Id_by_Staff(New):
    global count
    beginning_count=count

    for row in range(2,New.max_row+1):
        if New.sheet.cell(row=row,column=New.max_col+1).value!=1: #havenot been serched
            Find_Staff_ID(New,row)
    
    if not row%8:
            per=int(count*100/New.max_row)
            if (per<101):
                print("Filled row "+str(count)+"/"+str(New.max_row)+", "+str(per)+"%\r",end="",flush=True)

    print("Filled "+str(count-beginning_count)+" rows by the Workemail column")
    return count-beginning_count

def Find_Staff_ID(New,row):
    global count
    final_id=None

    staff_row=search_dict(New.linemanageremail_dict,New.sheet.cell(row=row,column=New.useremail_ID).value)
    if staff_row is not None: #found useremail in manager dictonary
        for i in staff_row:
            if New.sheet.cell(row=i,column=New.max_col+1).value==1: #staff has exact department id
                final_id=New.sheet.cell(row=row,column=New.userdepartmentid_ID).value
                break
        if final_id is None: #no department id exists among staff list
            for i in staff_row:
                final_id=Find_Staff_ID(New,i)
                if final_id is not None: #found department id in further staff
                    break
        if final_id is not None: #finally found department id after recusion
            count=count+1
            New.sheet.cell(row=row,column=New.max_col+1,value=1) #mark finish sign
            New.sheet.cell(row=row,column=New.userdepartmentid_ID,value=final_id) #print department id
            return final_id
    else:
        pass

def Copy_remaining_ID(New):
    global count
    copy_count=0
    zero_count=0

    for row in range(2,New.max_row+1):
        if New.sheet.cell(row=row,column=New.max_col+1).value!=1: #havenot been serched
            origin_id=New.sheet.cell(row=row,column=New.userdepartmentid_ID).value
            if origin_id!=1356:
                copy_count=copy_count+1
            else: #id equals 1356
                zero_count=zero_count+1
        
    print("Copyed "+str(copy_count)+" rows as their original department id")
    print("Rewriten "+str(zero_count)+" rows to 1356")
    print("Filled row "+str(New.max_row)+"/"+str(New.max_row)+", 100%\r")

#record dictionary (name in mapping sheet are all upper letters)
#input:
#   dict: dictionary
#   source: object sheet
#   name: name of sheet
#output:
#   None
def gen_dict(source,name):
    if name=="Tuser":
        source.create_dict("useremail")
        source.create_dict("linemanageremail")
        for row in range(2,source.max_row+1):
            #useremail_dict[useremail]=row
            key=source.sheet.cell(row=row,column=source.useremail_ID).value
            word=row
            append_dict(source.useremail_dict,key,word)

            #linemanageremail_dict[linemanageremail]=[row1,row2,...]
            key=source.sheet.cell(row=row,column=source.linemanageremail_ID).value
            if key is not None:
                if str(key).upper() not in source.linemanageremail_dict:
                    source.linemanageremail_dict[str(key).upper()]=[word]
                else:
                    source.linemanageremail_dict[str(key).upper()].append(word)

    if name=="EE":
        source.create_dict("workemail")
        for row in range(2,source.max_row+1):
            #workemail_dict[workemail]=row
            key=source.sheet.cell(row=row,column=source.workemail_ID).value
            word=row
            append_dict(source.workemail_dict,key,word)

    if name=="Mapping":
        source.create_dict("function3")
        for row in range(2,source.max_row+1):
            #function3_dict[function3]=departmentid
            key=source.sheet.cell(row=row,column=source.function3_ID).value
            word=source.sheet.cell(row=row,column=source.departmentid_ID).value
            append_dict(source.function3_dict,key,word)

#-----------------------------------------------------------------------------

#Start FSM
if __name__ == "__main__":
    #Variable initialization
    print("Start program\n\n")
    Tuser=Excel()
    EE=Excel()
    Mapping=Excel()
    Tuser.set_key_words(["useremail","userdepartmentid","usertype","linemanageremail"])
    EE.set_key_words(["function3","workemail"])
    Mapping.set_key_words(["function3","departmentid"])

    #FSM definition
    INIT=0
    MERGING=5
    FINISH=6
    EXIT=7

    state=INIT
    while 1:
        if state==INIT:  #Initialize window
            button=gui.buttonbox(msg="Welcome to Excel merging program.\n\n\n\
    Please import 3 documens in following order:\n\n\
        Tuser: "\
        +str(Tuser.add)+"\n\n\
        FTE+FA ESP: "\
        +str(EE.add)+"\n\n\
        Mapping:  "\
        +str(Mapping.add)\
            ,title="Excel Merging",choices=["Tuser","FTE+FA ESP","Mapping","Submit"])#button box return string
            if button=="Tuser":
                import_sheet(button,Tuser)
            elif button=="FTE+FA ESP":
                import_sheet("EE",EE)
            elif button=="Mapping":
                import_sheet(button,Mapping)
            elif button=="Submit":
                state=MERGING
            else:
                state=EXIT

        elif state==MERGING:
            if Tuser.add=="None" or EE.add=="None" or Mapping.add=="None": #missing file
                missing=[]
                if Tuser.add=="None":
                    missing.append("Tuser")
                if EE.add=="None":
                    missing.append("FTE and ESP")
                if Mapping.add=="None":
                    missing.append("Mapping")
                gui.msgbox(msg="Warning:\n\n\n\nMissing "+str(missing),title="Core file missing")
                state=INIT
            else:
                WB=Workbook()
                #create new workbook
                New=Excel() #new generated sheet
                New.set_key_words(Tuser.read_one_row(1,False))
                New.gen_new_sheet(WB=WB,sheet_id=0,name="Main")
                copy_sheet(Tuser,New)

                #generate dict
                gen_dict(New,"Tuser")
                gen_dict(EE,"EE")
                gen_dict(Mapping,"Mapping")

                New.sheet.cell(row=1,column=Tuser.sheet.max_column+1,value=1) #mark sign
                Search_EE_Sheet(New,EE,Mapping)
                zero_times=0
                for repeat_num in range(0,4):
                    manager_row=Find_Id_by_Manager(New)
                    if manager_row==0:
                        zero_times=zero_times+1
                    else:
                        zero_times=0
                    if zero_times==2:
                        break
                    staff_row=Find_Id_by_Staff(New)
                    if staff_row==0:
                        zero_times=zero_times+1
                    else:
                        zero_times=0
                    if zero_times==2:
                        break
                Copy_remaining_ID(New)

                for row in range(1,New.max_row+1):
                    New.sheet.cell(row=row,column=New.max_col+1,value="")

                print("\nStart saving...\n")

                state=FINISH

        elif state==FINISH:
            New.save_excel()
            if len(New.warning_list)>0:
                text=""
                for i in New.warning_list: #list out warning content
                    text+=(str(i)+"\n")
                gui.textbox(msg="Found "+str(len(New.warning_list))+" warnings:",text=text,title="Warning list")
            else:
                gui.msgbox("Finished")
            break
        
        elif state==EXIT:
            break

        else: 
            print("Error: no state finding!!!") #alart no state finding
            state=FINISH

    sys.exit(0)