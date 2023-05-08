#pack code: pyinstaller --name="test" --console --onefile --icon="C:\Users\86135\Desktop\Intern\Nancy\icons.ico" C:\Users\86135\Desktop\Intern\Nancy\fin_ver_2.py 

#import head file
import sys #import system file
import easygui as gui #import gui lib
from openpyxl import load_workbook #load existed sheet
from openpyxl import Workbook #create new sheet
from openpyxl.styles import PatternFill #fill cell's background color
import datetime #get date

#-----------------------------------------------------------------

class Excel(object):
    def __init__(self):
        self.add="None"
        self.warning_list=[]
        self.key_words=[]
        self.start_row=2

    #open an excel and also open one sheet after selecting
    def open_excel_and_sheet(self):
        self.WB=load_workbook(filename=str(self.add),data_only=True) #open workbook
        
        if len(self.WB.sheetnames)!=1:  #check if have multiple sheets
            self.sheet_name=gui.choicebox(msg="Multiple sheets founds. Please select one sheet below:",title="sheet selection",choices=self.WB.sheetnames)
            if self.sheet_name is None: #no choices
                self.add="None"
        else: #only have one sheet
            self.sheet_name=self.WB.sheetnames
            self.sheet_name=str(self.sheet_name)[2:len(self.sheet_name)-3] #remove [' '] inside original sheet name

        if self.sheet_name is not None: #Open seleted sheet and check head line
            self.sheet=self.WB[str(self.sheet_name)]
            self.max_row=self.sheet.max_row
            self.max_col=self.sheet.max_column
        
        if self.sheet_name is not None:
            print("Open file "+self.add+", "+self.sheet_name+"\n")

    #generate new sheet and create head line
    def gen_new_sheet(self,WB,sheet_id,name):
        self.WB=WB

        if len(WB.sheetnames)==1:
            self.sheet=WB.active
        else:
            self.sheet=WB.create_sheet(str(name),sheet_id)

        #create head line
        self.max_col=0
        self.max_row=1
        for i in self.key_words:
            self.max_col+=1 #shift column coordination
            self.sheet.cell(row=1,column=self.max_col,value=str(i))
            i=keep_letter_and_num(i).lower() #only keep number and letter
            exec("self.{}_ID = self.max_col".format(i)) #mark key words' column number. like: self.UserEmail_ID=5

    #enter key head words
    def set_key_words(self,keywords):
        for i in keywords:
            self.key_words.append(i)

    #get value of one row. head line mode will remove all " " , and turn to small letter
    def read_one_row(self,row,headLine):
        row_data=[]
        if headLine: #read head line
            for i in range(1,self.sheet.max_column+1):
                head_word=self.sheet.cell(row=row,column=i).value
                if head_word is not None:
                    head_word=keep_letter_and_num(head_word).lower() #only keep number and letter
                row_data.append(head_word)
        else:
            for i in range(1,self.sheet.max_column+1):
                row_data.append(self.sheet.cell(row=row,column=i).value)
        return row_data

    #check if have target word in head word
    def check_key_word(self,head_word,target_word):
        count=0
        for i in head_word:
            count=count+1
            if i==target_word:
                exec("self.{}_ID = count".format(i)) #mark key words' column number. like: self.UserEmail_ID=5
                return True #finds target
        print("Not found key word: "+str(target_word))
        return False

    #create new dictionary
    def create_dict(self,dict_name):
        exec("self.{}_dict={}".format(dict_name,{})) #create new dictionary. like self.email_dict={}

    # save excel
    def save_excel(self):
        date=datetime.datetime.now() #get system time
        save_add=gui.filesavebox(title="Save sheet",default="OpenAndFilled_"+date.strftime('%y%m%d')+".xlsx")
        if save_add is None:
            if gui.ynbox(msg="Warning. The worksheet has not been saved yet. \nPlease press YES if you want to save the worksheet."):
                print("Repeat the saving\n")
                self.save_excel()
            else:
                pass
        else:
            print("Saving sheet in "+str(save_add))
            self.WB.save(str(save_add))
            
#remove marks but keep letters and numbers in string(using ACSII form)
#input: 
#   msg:string
#output:
#   res:string
def keep_letter_and_num(msg):
    res=""
    for i in msg:
        acs_num=ord(i)
        if (acs_num<58 and acs_num>47) or (acs_num<123 and acs_num>96) or (acs_num<91 and acs_num>64):
            res+=i
    return res
#-----------------------------------------------------------------------------
def analyze_msg(excel,msg,row):
    column_array=[excel.obj1_ID,excel.kr11_ID,excel.kr12_ID,excel.kr13_ID,excel.obj2_ID,excel.kr21_ID,excel.kr22_ID,excel.kr23_ID,excel.obj3_ID,excel.kr31_ID,excel.kr32_ID,excel.kr33_ID]
    column_id=0
    last_line=""
    temp_msg=""

    msg_in_line=seperate_msg_into_line(msg)
    print(msg_in_line)

    for line in msg_in_line:
        #print(line)
        #print(str(ord(line[0])))
        #print(str(ord(line[1])))
        if ord(line[0])==49 and ord(line[1])==46: #found "1.", last line is obj
            if temp_msg=="":
                pass
            else:
                write_into_excel(excel,temp_msg,row,column_array[column_id])
                column_id+=1
            temp_msg=last_line
            last_line=""
            print("1")
        else: #last line is not obj
            temp_msg+=last_line
            print("2")

        if (ord(line[0])>47 and ord(line[0])<58) and ord(line[1])==46: #found "n."
            write_into_excel(excel,temp_msg,row,column_array[column_id])
            temp_msg=line[3:]
            column_id+=1
            print("3")
        elif ord(line[0])>64 and ord(line[0])<91: #first char in line is large letter
            last_line=line #not sure column id of this line
            print("last line: "+last_line)
            print("4")
        else:
            temp_msg+=line
            print("5")
    
    write_into_excel(excel,temp_msg,row,column_array[column_id]) #record last line

def write_into_excel(excel,msg,row,column):
    print(str(column)+" "+msg)
    #for char in msg:
        #print(ord(char))
    excel.sheet.cell(row=row,column=column,value=msg)

def seperate_msg_into_line(msg):
    temp_msg=""
    msg_in_line=[]

    for char in msg:
        if ord(char)==10:
            msg_in_line.append(temp_msg)
            temp_msg=""
        elif ord(char)>-1 and ord(char)<32:
            pass
        else:
            temp_msg+=char

    msg_in_line.append(temp_msg) #record last line
    return msg_in_line

#-----------------------------------------------------------------------------

#Start FSM
if __name__ == "__main__":
    #Variable initialization
    print("Start program\n\n")

    WB=Workbook()

    #create new workbook
    New=Excel() #new generated sheet
    New.set_key_words(["EMPID","UserEmail","Year","Obj1","Obj1ID","Obj1AlignID","Obj1AlignUserEmail","KR1.1","KR1.1ID","KR1.2","KR1.2ID","KR1.3","KR1.3ID","Obj2","Obj2ID","Obj2AlignID","Obj2AlignUserEmail","KR2.1","KR2.1ID","KR2.2","KR2.2ID","KR2.3","KR2.3ID","Obj3","Obj3ID","Obj3AlignID","Obj3AlignUserEmail","KR3.1","KR3.1ID","KR3.2","KR3.2ID","KR3.3","KR3.3ID"])
    New.gen_new_sheet(WB=WB,sheet_id=0,name="Main")

    yellow_color=PatternFill('solid',fgColor='ffff00')
    row=2

    #FSM definition
    state="INIT"
    while 1:
        if state=="INIT":  #Initialize window
            choice=gui.buttonbox(msg="Message",title="OKR",choices=["Import New Object","Output as Execl"]) 

            if choice=="Import New Object":
                state="IMPORT"
            elif choice=="Output as Execl":
                state="FINISH"
            else:
                state="EXIT"

        elif state=="IMPORT":
            msg=gui.textbox(msg='Copy contents of PDF into the text box below:',title='Import content',text='')

            if msg is None:
                pass
            else:
                analyze_msg(New,msg,row)
                row+=1

            state="INIT"

        elif state=="FINISH":
            New.save_excel()
            if len(New.warning_list)>0:
                text=""
                for i in New.warning_list: #list out warning content
                    text+=(str(i)+"\n")
                gui.textbox(msg="Found "+str(len(New.warning_list))+" warnings:",text=text,title="Warning list")
            else:
                gui.msgbox("Finished")
            break
        
        elif state=="EXIT":
            break

        else: 
            print("Error: no state finding!!!") #alart no state finding
            state="FINISH"

    sys.exit(0)